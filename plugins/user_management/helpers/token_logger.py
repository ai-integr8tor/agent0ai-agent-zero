import io
from usr.plugins.user_management.helpers.db import execute_query, execute_write


def log_tokens(user_id, context_id, model, input_tokens=0, output_tokens=0):
    """Insert a token usage record."""
    execute_write(
        "INSERT INTO um_token_usage "
        "(user_id, context_id, model, input_tokens, output_tokens) "
        "VALUES (%s, %s, %s, %s, %s)",
        (user_id, context_id, model, input_tokens, output_tokens),
    )


def get_usage(
    user_id=None, from_date=None, to_date=None,
    model=None, context_id=None, limit=1000,
):
    """Query token usage with optional filters."""
    conds, params = [], []
    if user_id is not None:
        conds.append("tu.user_id = %s"); params.append(user_id)
    if from_date:
        conds.append("tu.timestamp >= %s"); params.append(from_date)
    if to_date:
        conds.append("tu.timestamp <= %s"); params.append(to_date)
    if model:
        conds.append("tu.model = %s"); params.append(model)
    if context_id:
        conds.append("tu.context_id = %s"); params.append(context_id)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    params.append(limit)
    return execute_query(
        f"""SELECT tu.id, tu.user_id, u.username, tu.context_id, tu.model,
                   tu.input_tokens, tu.output_tokens, tu.timestamp
            FROM um_token_usage tu
            LEFT JOIN um_users u ON tu.user_id = u.id
            {where}
            ORDER BY tu.timestamp DESC
            LIMIT %s""",
        params,
    )


def get_usage_summary(user_id=None, group_by="day", from_date=None, to_date=None):
    """Aggregate token usage grouped by day, user, or model."""
    conds, params = [], []
    if user_id is not None:
        conds.append("tu.user_id = %s"); params.append(user_id)
    if from_date:
        conds.append("tu.timestamp >= %s"); params.append(from_date)
    if to_date:
        conds.append("tu.timestamp <= %s"); params.append(to_date)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    if group_by == "user":
        gcol, scol = "u.username", "u.username AS group_key"
    elif group_by == "model":
        gcol, scol = "tu.model", "tu.model AS group_key"
    else:
        gcol, scol = "DATE(tu.timestamp)", "DATE(tu.timestamp)::text AS group_key"

    return execute_query(
        f"""SELECT {scol},
                   COUNT(*) AS request_count,
                   COALESCE(SUM(tu.input_tokens), 0) AS total_input_tokens,
                   COALESCE(SUM(tu.output_tokens), 0) AS total_output_tokens,
                   COALESCE(SUM(tu.input_tokens + tu.output_tokens), 0) AS total_tokens
            FROM um_token_usage tu
            LEFT JOIN um_users u ON tu.user_id = u.id
            {where}
            GROUP BY {gcol}
            ORDER BY {gcol}""",
        params,
    )


def export_to_excel(user_id=None, from_date=None, to_date=None):
    """Generate an Excel workbook with summary + detail sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    data = get_usage(user_id=user_id, from_date=from_date, to_date=to_date, limit=50000)
    summary = get_usage_summary(
        user_id=user_id, from_date=from_date, to_date=to_date, group_by="day"
    )

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    for ci, h in enumerate(
        ["Date", "Requests", "Input Tokens", "Output Tokens", "Total Tokens"], 1
    ):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill = hf, hfill
    for ri, row in enumerate(summary, 2):
        ws.cell(row=ri, column=1, value=str(row["group_key"]))
        ws.cell(row=ri, column=2, value=row["request_count"])
        ws.cell(row=ri, column=3, value=row["total_input_tokens"])
        ws.cell(row=ri, column=4, value=row["total_output_tokens"])
        ws.cell(row=ri, column=5, value=row["total_tokens"])
    for ci in range(1, 6):
        ws.column_dimensions[chr(64 + ci)].width = 18

    # --- Detail sheet ---
    ws2 = wb.create_sheet("Details")
    for ci, h in enumerate(
        ["ID", "Username", "Context", "Model", "Input Tokens", "Output Tokens", "Timestamp"], 1
    ):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font, c.fill = hf, hfill
    for ri, row in enumerate(data, 2):
        ws2.cell(row=ri, column=1, value=row["id"])
        ws2.cell(row=ri, column=2, value=row.get("username") or "N/A")
        ws2.cell(row=ri, column=3, value=row["context_id"])
        ws2.cell(row=ri, column=4, value=row["model"])
        ws2.cell(row=ri, column=5, value=row["input_tokens"])
        ws2.cell(row=ri, column=6, value=row["output_tokens"])
        ws2.cell(row=ri, column=7, value=str(row["timestamp"]))
    for ci in range(1, 8):
        ws2.column_dimensions[chr(64 + ci)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
